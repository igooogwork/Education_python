# Импортируем библиотеку openpyxl для работы с Excel-файлами
import openpyxl

# Пути к файлам
excel_file = 'C:\\Users\\lawrence\\Desktop\\baza1.xlsx'  # Путь к исходному файлу Excel
result_file = 'C:\\Users\\lawrence\\Desktop\\result.xlsx'  # Путь к файлу, в который будут сохранены результаты

# Названия листов
source_sheet_name = 'Продажа'  # Имя листа в исходном файле
new_sheet_name = 'Продажа'  # Имя листа для нового файла

# Определение соответствия между названиями столбцов и их индексами
columns = {'address': 1, 'floor': 3, 'type': 4, 'Square': 5, 'price': 6}

# Заголовки столбцов для нового листа
result_sheet_header = ['Адрес', 'Этаж', 'Тип', 'Площадь', 'Цена']

# Загрузка исходного файла Excel в режиме "только для чтения"
workbook = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)

# Выбор листа для обработки
sheet = workbook[source_sheet_name]

# Создание нового файла Excel
new_workbook = openpyxl.Workbook()

# Создание нового листа с заданным именем
new_sheet = new_workbook.create_sheet(title=new_sheet_name)

# Добавление заголовков столбцов в новый лист
new_sheet.append(result_sheet_header)

# Обработка данных из исходного листа
for row in sheet.iter_rows(min_row=2, values_only=True):
    # Создание словаря с данными из строки
    data_dict = {key: row[col_index - 1] for key, col_index in columns.items()}

    # Проверка, что хотя бы одно значение в строке не является None
    if any(data_dict.values()):
        # Добавление строки с данными в новый лист
        new_sheet.append([data_dict[key] for key in columns])

# Сохранение нового файла Excel с результатами
new_workbook.save(result_file)

# Вывод сообщения о завершении операции
print("Данные сохранены в result.xlsx")
